#!/usr/bin/env python3
"""
Create one unified Excel table from all scraper outputs.

Stacks rows from each source institution into a single table,
consolidates column names, cleans numeric values, and applies Excel formatting.
"""
import re
import sys
import os
import json
from datetime import datetime
from pathlib import Path

import pandas as pd
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


DATE_RE = re.compile(r"(\d{4}-\d{2}-\d{2})")
FUND_AGE_BUCKETS = [
    "2003-2009",
    "1996-2002",
    "1989-1995",
    "1982-1988",
    "1975-1981",
    "1968-1974",
    "1961-1967",
    "1954-1960",
]
PROVIDER_ORDER = ["allianz", "artea", "goindex", "luminor", "seb", "swedbank"]
PROVIDER_ORDER_MAP = {name: idx for idx, name in enumerate(PROVIDER_ORDER)}
FUND_BUCKET_MAP = {bucket: idx for idx, bucket in enumerate(FUND_AGE_BUCKETS)}


def collect_report_files(docs_dir: Path, allowed_dates=None, max_date=None):
    reports = []
    for path in sorted(docs_dir.glob("pension_data_combined_*.html")):
        m = DATE_RE.search(path.name)
        if not m:
            continue
        report_date = m.group(1)
        if allowed_dates is not None and report_date not in allowed_dates:
            continue
        if max_date is not None and report_date > max_date:
            continue
        reports.append({
            "date": report_date,
            "html": path.name,
            "xlsx": f"pension_data_combined_{report_date}.xlsx",
        })
    return sorted(reports, key=lambda r: r["date"])


def discover_complete_snapshot_dates():
    """Return all dates that exist across all required providers."""
    dates_by_institution = {}

    for path in Path(".").glob("*.xlsx"):
        lower = path.name.lower()
        if path.name.startswith("~$"):
            continue
        if "combined" in lower:
            continue
        if not DATE_RE.search(path.name):
            continue

        source_name, file_date = parse_source_and_date(path.name)
        if not source_name or not file_date:
            continue

        institution = institution_from_source(source_name)
        dates_by_institution.setdefault(institution, set()).add(file_date)

    required_providers = set(PROVIDER_ORDER)
    if not required_providers.issubset(dates_by_institution.keys()):
        return set()

    common_dates = None
    for provider in PROVIDER_ORDER:
        provider_dates = dates_by_institution.get(provider, set())
        common_dates = provider_dates if common_dates is None else (common_dates & provider_dates)

    return common_dates or set()


def discover_provider_floor_date():
    """
    Return the minimum of latest provider dates.

    Example:
    - providers latest: [2026-06-02, 2026-06-02, 2026-06-03, ...]
    - floor date: 2026-06-02
    """
    latest_by_provider = {}

    for path in Path(".").glob("*.xlsx"):
        lower = path.name.lower()
        if path.name.startswith("~$"):
            continue
        if "combined" in lower:
            continue
        if not DATE_RE.search(path.name):
            continue

        source_name, file_date = parse_source_and_date(path.name)
        if not source_name or not file_date:
            continue

        provider = institution_from_source(source_name)
        current = latest_by_provider.get(provider)
        if current is None or file_date > current:
            latest_by_provider[provider] = file_date

    required = set(PROVIDER_ORDER)
    if not required.issubset(latest_by_provider.keys()):
        return None

    return min(latest_by_provider[p] for p in PROVIDER_ORDER)


def prune_old_reports(docs_dir: Path, max_keep: int = 14):
    reports = collect_report_files(docs_dir)
    if len(reports) <= max_keep:
        return

    old_reports = reports[:-max_keep]
    for report in old_reports:
        for report_file in (report["html"], report["xlsx"]):
            path = docs_dir / report_file
            if path.exists():
                try:
                    path.unlink()
                except Exception:
                    pass

    # Also remove older root XLSX copies from the repository root.
    xlsx_files = [path for path in Path(".").glob("pension_data_combined_*.xlsx") if DATE_RE.search(path.name)]
    xlsx_files.sort(key=lambda p: DATE_RE.search(p.name).group(1))
    for path in xlsx_files[:-max_keep]:
        if path.exists():
            try:
                path.unlink()
            except Exception:
                pass


def format_report_index(reports, trigger_api_url):
    if not reports:
        return (
            "<!doctype html>\n"
            "<html lang=\"en\"><head><meta charset=\"utf-8\"><meta name=\"viewport\" content=\"width=device-width,initial-scale=1,viewport-fit=cover\">\n"
            "<title>Daily pension reports</title>\n"
            "<script src=\"https://cdn.tailwindcss.com\"></script>\n"
            "</head><body class=\"bg-slate-50 text-slate-900 antialiased\">\n"
            "<div class=\"min-h-screen flex items-center justify-center px-4 py-8 pb-[max(2rem,env(safe-area-inset-bottom))]\">\n"
            "  <main class=\"w-full max-w-4xl\">\n"
            "    <section class=\"rounded-[32px] border border-slate-200 bg-white/95 p-8 shadow-[0_30px_80px_-40px_rgba(15,23,42,0.35)] backdrop-blur-xl\">\n"
            "      <div class=\"flex flex-col gap-6 md:gap-8\">\n"
            "        <div class=\"flex flex-col gap-4 md:flex-row md:items-end md:justify-between\">\n"
            "          <div>\n"
            "            <p class=\"text-sm uppercase tracking-[0.3em] text-slate-400\">Modern fund archive</p>\n"
            "            <h1 class=\"mt-2 text-3xl font-semibold tracking-tight text-slate-950 sm:text-4xl\">Daily pension reports</h1>\n"
            "            <p class=\"mt-3 max-w-2xl text-base leading-7 text-slate-600\">No reports have been generated yet. Check back after the first run.</p>\n"
            "          </div>\n"
            "        </div>\n"
            "      </div>\n"
            "    </section>\n"
            "  </main>\n"
            "</div>\n"
            "</body></html>"
        )

    reports_js = ",\n        ".join(
        [
            f'{{date: "{r["date"]}", html: "{r["html"]}", xlsx: "{r["xlsx"]}"}}'
            for r in reports
        ]
    )
    latest = reports[-1]

    return (
        "<!doctype html>\n"
        "<html lang=\"en\"><head><meta charset=\"utf-8\"><meta name=\"viewport\" content=\"width=device-width,initial-scale=1,viewport-fit=cover\">\n"
        "<title>Daily pension data</title>\n"
        "<script src=\"https://cdn.tailwindcss.com\"></script>\n"
        "</head><body class=\"bg-slate-50 text-slate-900 antialiased\">\n"
        "<div class=\"min-h-screen px-3 py-4 pb-[max(2rem,env(safe-area-inset-bottom))] sm:px-6 sm:py-8 lg:px-8\">\n"
        "  <main class=\"mx-auto w-full max-w-7xl\">\n"
        "    <section class=\"rounded-[24px] sm:rounded-[32px] border border-slate-200 bg-white/95 p-4 sm:p-6 shadow-[0_30px_80px_-40px_rgba(15,23,42,0.35)]\">\n"
        "      <div class=\"flex flex-col gap-6\">\n"
        "        <div class=\"flex flex-col gap-4 md:flex-row md:items-end md:justify-between\">\n"
        "          <div>\n"
        "            <p class=\"text-sm uppercase tracking-[0.3em] text-slate-400\">Daily pension funds</p>\n"
        "            <h1 class=\"mt-2 text-3xl font-semibold tracking-tight text-slate-950 sm:text-4xl\">Single-page dashboard</h1>\n"
        "            <p class=\"mt-3 max-w-2xl text-base leading-7 text-slate-600\">Choose a date and view the full report below without leaving this page.</p>\n"
        "          </div>\n"
        "          <div class=\"w-full md:w-auto grid grid-cols-1 sm:grid-cols-[auto_minmax(12rem,1fr)_auto_auto] items-center gap-2 sm:gap-3\">\n"
        "            <label class=\"text-sm font-medium text-slate-700\" for=\"report-select\">Report date</label>\n"
        "            <select id=\"report-select\" class=\"w-full rounded-full border border-slate-300 bg-slate-50 px-4 py-2.5 text-sm font-semibold text-slate-700 shadow-sm outline-none transition focus:border-slate-500 focus:ring-2 focus:ring-slate-200\"></select>\n"
        "            <a id=\"download-link\" class=\"w-full sm:w-auto text-center rounded-full bg-slate-900 px-4 py-2.5 text-sm font-semibold text-white transition hover:bg-slate-700\" href=\"#\" download>Download Excel</a>\n"
        "          </div>\n"
        "        </div>\n"
        "        <div class=\"rounded-3xl border border-slate-200 bg-slate-50/80 p-4 text-sm text-slate-600\">\n"
                "          Device mode is enabled: reports generated from this browser are shown only on this browser.\n"
        "          <p class=\"mt-2 text-xs text-slate-500\">Automatic checks run at 15:00, 16:00, 17:00, and 18:00 Vilnius time until new data appears.</p>\n"
        "          <p id=\"last-check-note\" class=\"mt-1 text-xs text-slate-500\">Last automatic check: not yet</p>\n"
        "        </div>\n"
        "        <div id=\"report-body\" class=\"overflow-x-auto rounded-3xl border border-slate-200 bg-white min-h-[16rem]\">\n"
        "          <div id=\"report-content\"><p class=\"text-center text-slate-400 text-sm py-16 px-4\">Published reports will sync automatically on this browser.</p></div>\n"
        "        </div>\n"
        "      </div>\n"
        "    </section>\n"
        "  </main>\n"
        "</div>\n"
        "<script>\n"
        "const reports = [\n"
        + reports_js +
        "\n];\n"
        f"const TRIGGER_API_URL = '{trigger_api_url}';\n"
        "const DEVICE_ID_KEY = 'ltFundsDeviceId';\n"
        "const DEVICE_DATES_PREFIX = 'ltFundsVisibleDates:';\n"
        "const PENDING_UPDATE_KEY = 'ltFundsPendingBaseline';\n"
        "const LAST_PUBLISHED_CHECK_KEY = 'ltFundsLastPublishedCheckAt';\n"
        "const AUTO_CHECK_STATE_KEY = 'ltFundsAutoCheckState';\n"
        "const VILNIUS_TIME_ZONE = 'Europe/Vilnius';\n"
        "const AUTO_CHECK_SLOTS = [15 * 60, 16 * 60, 17 * 60, 18 * 60];\n"
        "let autoCheckTimerId = null;\n"
        "function byNewest(a, b) { return b.date.localeCompare(a.date); }\n"
        "function getDeviceId() {\n"
        "  let id = localStorage.getItem(DEVICE_ID_KEY);\n"
        "  if (id) return id;\n"
        "  if (window.crypto && window.crypto.randomUUID) {\n"
        "    id = window.crypto.randomUUID();\n"
        "  } else {\n"
        "    id = 'dev-' + Date.now() + '-' + Math.random().toString(16).slice(2);\n"
        "  }\n"
        "  localStorage.setItem(DEVICE_ID_KEY, id);\n"
        "  return id;\n"
        "}\n"
        "function visibleDatesKey() {\n"
        "  return DEVICE_DATES_PREFIX + getDeviceId();\n"
        "}\n"
        "function getGlobalSortedReports() {\n"
        "  return reports.slice().sort(byNewest);\n"
        "}\n"
        "function readVisibleDates() {\n"
        "  const key = visibleDatesKey();\n"
        "  const raw = localStorage.getItem(key);\n"
        "  if (raw) {\n"
        "    try {\n"
        "      const parsed = JSON.parse(raw);\n"
        "      if (Array.isArray(parsed)) return parsed;\n"
        "    } catch (e) {}\n"
        "  }\n"
        "  localStorage.setItem(key, JSON.stringify([]));\n"
        "  return [];\n"
        "}\n"
        "function saveVisibleDates(dates) {\n"
        "  const uniq = Array.from(new Set(dates)).sort();\n"
        "  localStorage.setItem(visibleDatesKey(), JSON.stringify(uniq));\n"
        "}\n"
        "function deviceReports() {\n"
        "  const allowed = new Set(readVisibleDates());\n"
        "  return getGlobalSortedReports().filter(function(r){ return allowed.has(r.date); });\n"
        "}\n"
        "function formatLocalTimestamp(iso) {\n"
        "  if (!iso) return 'not yet';\n"
        "  const d = new Date(iso);\n"
        "  if (Number.isNaN(d.getTime())) return 'not yet';\n"
        "  return d.toLocaleString();\n"
        "}\n"
        "function updateLastCheckLabel() {\n"
        "  const el = document.getElementById('last-check-note');\n"
        "  if (!el) return;\n"
        "  // Try to fetch server-side last run timestamp first\n"
        "  if (window.location.protocol !== 'file:') {\n"
        "    fetch('last_run.json?t=' + Date.now(), { cache: 'no-store' })\n"
        "      .then(function(r){ return r.ok ? r.json() : null; })\n"
        "      .then(function(data){\n"
        "        if (data && data.lastRunAt) {\n"
        "          el.textContent = 'Last pipeline run: ' + formatLocalTimestamp(data.lastRunAt);\n"
        "        } else {\n"
        "          const last = localStorage.getItem(LAST_PUBLISHED_CHECK_KEY);\n"
        "          el.textContent = 'Last automatic check: ' + formatLocalTimestamp(last);\n"
        "        }\n"
        "      })\n"
        "      .catch(function(){\n"
        "        const last = localStorage.getItem(LAST_PUBLISHED_CHECK_KEY);\n"
        "        el.textContent = 'Last automatic check: ' + formatLocalTimestamp(last);\n"
        "      });\n"
        "  } else {\n"
        "    const last = localStorage.getItem(LAST_PUBLISHED_CHECK_KEY);\n"
        "    el.textContent = 'Last automatic check: ' + formatLocalTimestamp(last);\n"
        "  }\n"
        "}\n"
        "function getVilniusNowParts() {\n"
        "  const parts = new Intl.DateTimeFormat('en-CA', {\n"
        "    timeZone: VILNIUS_TIME_ZONE,\n"
        "    year: 'numeric',\n"
        "    month: '2-digit',\n"
        "    day: '2-digit',\n"
        "    hour: '2-digit',\n"
        "    minute: '2-digit',\n"
        "    second: '2-digit',\n"
        "    hourCycle: 'h23'\n"
        "  }).formatToParts(new Date()).reduce(function(acc, part) {\n"
        "    if (part.type !== 'literal') acc[part.type] = part.value;\n"
        "    return acc;\n"
        "  }, {});\n"
        "  const hour = Number(parts.hour || '0');\n"
        "  const minute = Number(parts.minute || '0');\n"
        "  const second = Number(parts.second || '0');\n"
        "  return {\n"
        "    dateKey: [parts.year || '0000', parts.month || '00', parts.day || '00'].join('-'),\n"
        "    totalMinutes: hour * 60 + minute,\n"
        "    totalSeconds: hour * 3600 + minute * 60 + second\n"
        "  };\n"
        "}\n"
        "function readAutoCheckState(dateKey) {\n"
        "  const raw = localStorage.getItem(AUTO_CHECK_STATE_KEY);\n"
        "  if (!raw) return { dateKey: dateKey, completedSlots: [], found: false };\n"
        "  try {\n"
        "    const parsed = JSON.parse(raw);\n"
        "    if (parsed && parsed.dateKey === dateKey && Array.isArray(parsed.completedSlots)) {\n"
        "      return {\n"
        "        dateKey: parsed.dateKey,\n"
        "        completedSlots: parsed.completedSlots,\n"
        "        found: Boolean(parsed.found)\n"
        "      };\n"
        "    }\n"
        "  } catch (e) {}\n"
        "  return { dateKey: dateKey, completedSlots: [], found: false };\n"
        "}\n"
        "function writeAutoCheckState(state) {\n"
        "  localStorage.setItem(AUTO_CHECK_STATE_KEY, JSON.stringify({\n"
        "    dateKey: state.dateKey,\n"
        "    completedSlots: state.completedSlots,\n"
        "    found: Boolean(state.found)\n"
        "  }));\n"
        "}\n"
        "function currentEligibleSlot(totalMinutes) {\n"
        "  for (let i = AUTO_CHECK_SLOTS.length - 1; i >= 0; i -= 1) {\n"
        "    if (totalMinutes >= AUTO_CHECK_SLOTS[i]) return AUTO_CHECK_SLOTS[i];\n"
        "  }\n"
        "  return null;\n"
        "}\n"
        "function computeNextAutoCheckDelayMs() {\n"
        "  const now = getVilniusNowParts();\n"
        "  for (let i = 0; i < AUTO_CHECK_SLOTS.length; i += 1) {\n"
        "    const slotMinutes = AUTO_CHECK_SLOTS[i];\n"
        "    const slotSeconds = slotMinutes * 60;\n"
        "    if (now.totalSeconds < slotSeconds) {\n"
        "      return Math.max(1000, (slotSeconds - now.totalSeconds) * 1000);\n"
        "    }\n"
        "  }\n"
        "  return Math.max(1000, ((24 * 60 + AUTO_CHECK_SLOTS[0]) * 60 - now.totalSeconds) * 1000);\n"
        "}\n"
        "function markPublishedCheckNow() {\n"
        "  const nowIso = new Date().toISOString();\n"
        "  localStorage.setItem(LAST_PUBLISHED_CHECK_KEY, nowIso);\n"
        "  updateLastCheckLabel();\n"
        "  return nowIso;\n"
        "}\n"
        "function parseReportsFromIndexHtml(html) {\n"
        "  const match = html.match(/const reports = \\[([\\s\\S]*?)\\n\\];/);\n"
        "  if (!match || !match[1]) return [];\n"
        "  try {\n"
        "    const parsed = (new Function('return [' + match[1] + '];'))();\n"
        "    if (!Array.isArray(parsed)) return [];\n"
        "    return parsed.filter(function(item){ return item && item.date && item.html && item.xlsx; });\n"
        "  } catch (e) {\n"
        "    return [];\n"
        "  }\n"
        "}\n"
        "async function fetchPublishedReports() {\n"
        "  const pagePath = window.location.pathname || '/';\n"
        "  const res = await fetch(pagePath + '?t=' + Date.now(), { cache: 'no-store' });\n"
        "  if (!res.ok) throw new Error('publish check failed (' + res.status + ')');\n"
        "  const html = await res.text();\n"
        "  return parseReportsFromIndexHtml(html);\n"
        "}\n"
        "function mergePublishedReportsIntoVisible(published) {\n"
        "  if (!published || !published.length) return 0;\n"
        "  const visibleSet = new Set(readVisibleDates());\n"
        "  let added = 0;\n"
        "  published.forEach(function(report){\n"
        "    if (!visibleSet.has(report.date)) {\n"
        "      visibleSet.add(report.date);\n"
        "      added += 1;\n"
        "    }\n"
        "  });\n"
        "  if (added) saveVisibleDates(Array.from(visibleSet));\n"
        "  return added;\n"
        "}\n"
        "function scheduleNextAutoCheck() {\n"
        "  if (autoCheckTimerId) window.clearTimeout(autoCheckTimerId);\n"
        "  const now = getVilniusNowParts();\n"
        "  const state = readAutoCheckState(now.dateKey);\n"
        "  if (state.found) return;\n"
        "  autoCheckTimerId = window.setTimeout(function() {\n"
        "    maybeAutoCheckPublishedReports();\n"
        "  }, computeNextAutoCheckDelayMs());\n"
        "}\n"
        "async function maybeAutoCheckPublishedReports() {\n"
        "  if (window.location.protocol === 'file:') {\n"
        "    updateLastCheckLabel();\n"
        "    return;\n"
        "  }\n"
        "  const now = getVilniusNowParts();\n"
        "  const slot = currentEligibleSlot(now.totalMinutes);\n"
        "  const state = readAutoCheckState(now.dateKey);\n"
        "  if (state.found) {\n"
        "    updateLastCheckLabel();\n"
        "    return;\n"
        "  }\n"
        "  if (slot === null || state.completedSlots.indexOf(slot) !== -1) {\n"
        "    updateLastCheckLabel();\n"
        "    scheduleNextAutoCheck();\n"
        "    return;\n"
        "  }\n"
        "  markPublishedCheckNow();\n"
        "  state.completedSlots.push(slot);\n"
        "  writeAutoCheckState(state);\n"
        "  try {\n"
        "    const published = (await fetchPublishedReports()).sort(byNewest);\n"
        "    if (!published.length) return;\n"
        "    let added = mergePublishedReportsIntoVisible(published);\n"
        "    if (!added) return;\n"
        "    state.found = true;\n"
        "    writeAutoCheckState(state);\n"
        "    showNotice('New published data found (' + added + '). Reloading dashboard...', 'success');\n"
        "    window.location.reload();\n"
        "  } catch (e) {\n"
        "    // Keep silent here: managed devices may block some requests.\n"
        "  } finally {\n"
        "    if (!state.found) scheduleNextAutoCheck();\n"
        "  }\n"
        "}\n"
        "async function setSelection(report) {\n"
        "  const link = document.getElementById('download-link');\n"
        "  const body = document.getElementById('report-content');\n"
        "  link.href = report.xlsx;\n"
        "  body.innerHTML = '<p class=\"text-slate-400 text-sm text-center py-16 px-4\">Loading report\u2026</p>';\n"
        "  try {\n"
        "    const res = await fetch(report.html + '?t=' + Date.now(), { cache: 'no-store' });\n"
        "    const html = await res.text();\n"
        "    const tmp = document.createElement('div');\n"
        "    const m = html.match(/<body[^>]*>([\\s\\S]*)<\\/body>/i);\n"
        "    tmp.innerHTML = m ? m[1] : html;\n"
        "    const style = tmp.querySelector('style');\n"
        "    const table = tmp.querySelector('table');\n"
        "    body.innerHTML = '';\n"
        "    if (style) body.appendChild(style.cloneNode(true));\n"
        "    if (table) body.appendChild(table.cloneNode(true));\n"
        "    else body.innerHTML = '<p class=\"text-red-400 text-sm text-center py-16\">Could not extract table.</p>';\n"
        "  } catch(e) {\n"
        "    body.innerHTML = '<p class=\"text-red-500 text-sm text-center py-16\">Failed to load report: ' + e.message + '</p>';\n"
        "  }\n"
        "}\n"
        "function showNotice(message, type) {\n"
        "  const bgClass = type === 'success' ? 'bg-green-50 border-green-200 text-green-800' :\n"
        "                  type === 'warning' ? 'bg-yellow-50 border-yellow-200 text-yellow-800' :\n"
        "                  type === 'error' ? 'bg-red-50 border-red-200 text-red-800' :\n"
        "                  'bg-blue-50 border-blue-200 text-blue-800';\n"
        "  const n = document.createElement('div');\n"
        "  n.className = 'fixed top-3 left-3 right-3 sm:left-auto sm:right-4 sm:max-w-sm p-4 rounded-2xl border ' + bgClass + ' shadow-lg z-50';\n"
        "  n.textContent = message;\n"
        "  document.body.appendChild(n);\n"
        "  setTimeout(function(){ n.remove(); }, 7000);\n"
        "}\n"
        "function getStatusUrl(afterIso) {\n"
        "  const base = TRIGGER_API_URL.endsWith('/trigger')\n"
        "    ? TRIGGER_API_URL.slice(0, -8) + '/status'\n"
        "    : TRIGGER_API_URL.replace(/\\/+$/, '') + '/status';\n"
        "  return base + '?after=' + encodeURIComponent(afterIso || '');\n"
        "}\n"
        "async function triggerWorkflow() {\n"
        "  if (!TRIGGER_API_URL) {\n"
        "    showNotice('Trigger endpoint is not configured.', 'error');\n"
        "    return null;\n"
        "  }\n"
        "  if (window.location.protocol === 'file:') {\n"
        "    showNotice('Update trigger is unavailable from local file preview. Open the deployed page or run a local web server.', 'warning');\n"
        "    return null;\n"
        "  }\n"
        "  try {\n"
        "    const res = await fetch(TRIGGER_API_URL, {\n"
        "      method: 'POST',\n"
        "      headers: { 'Content-Type': 'application/json' },\n"
        "      body: JSON.stringify({ action: 'trigger_daily_publish' })\n"
        "    });\n"
        "    const payload = await res.json().catch(function(){ return {}; });\n"
        "    if (res.ok) return payload.dispatchedAt || new Date().toISOString();\n"
        "    const detail = payload && payload.error ? ': ' + payload.error : '';\n"
        "    showNotice('Trigger failed (' + res.status + ')' + detail, 'error');\n"
        "    return null;\n"
        "  } catch (e) {\n"
        "    var msg = (e && e.message) ? e.message : 'network error';\n"
        "    if (window.location.protocol === 'file:') {\n"
        "      msg = 'Request blocked in this context. Use the deployed page or local web server.';\n"
        "    } else if (msg === 'Failed to fetch') {\n"
        "      msg = 'Network error. Check your connection, or it may be a CORS restriction. Check browser console (F12) for details.';\n"
        "    }\n"
        "    showNotice('Error calling trigger endpoint: ' + msg, 'error');\n"
        "    return null;\n"
        "  }\n"
        "}\n"
        "async function waitForBuildCompletion(dispatchedAt, maxWaitMs) {\n"
        "  maxWaitMs = maxWaitMs || 420000;\n"
        "  const start = Date.now();\n"
        "  const statusUrl = getStatusUrl(dispatchedAt);\n"
        "  while (Date.now() - start < maxWaitMs) {\n"
        "    await new Promise(function(resolve){ setTimeout(resolve, 8000); });\n"
        "    try {\n"
        "      const res = await fetch(statusUrl + '&t=' + Date.now(), { cache: 'no-store' });\n"
        "      if (!res.ok) continue;\n"
        "      const payload = await res.json().catch(function(){ return {}; });\n"
        "      if (payload.run && payload.run.status === 'completed') return payload.run;\n"
        "    } catch (e) {}\n"
        "  }\n"
        "  return null;\n"
        "}\n"
        "async function initialPublishedSync() {\n"
        "  try {\n"
        "    const published = (await fetchPublishedReports()).sort(byNewest);\n"
        "    if (!published.length) return;\n"
        "    const added = mergePublishedReportsIntoVisible(published);\n"
        "    if (!added) return;\n"
        "    const now = getVilniusNowParts();\n"
        "    const state = readAutoCheckState(now.dateKey);\n"
        "    state.found = true;\n"
        "    writeAutoCheckState(state);\n"
        "    showNotice('Loaded ' + added + ' newly published report(s). Reloading dashboard...', 'success');\n"
        "    window.location.reload();\n"
        "  } catch (e) {\n"
        "    // Silent on managed devices when networking is restricted.\n"
        "  }\n"
        "}\n"
        "function initIndex() {\n"
        "  const select = document.getElementById('report-select');\n"
        "  updateLastCheckLabel();\n"
        "  // Pending post-build check: see if CI produced a new date.\n"
        "  const pending = sessionStorage.getItem(PENDING_UPDATE_KEY);\n"
        "  if (pending) {\n"
        "    sessionStorage.removeItem(PENDING_UPDATE_KEY);\n"
        "    const newestGlobal = getGlobalSortedReports()[0];\n"
        "    if (newestGlobal && newestGlobal.date > pending) {\n"
        "      const dates = readVisibleDates();\n"
        "      dates.push(newestGlobal.date);\n"
        "      saveVisibleDates(dates);\n"
        "      showNotice('New report available: ' + newestGlobal.date, 'success');\n"
        "    } else {\n"
        "      showNotice('Already up to date — no newer report after this pipeline run.', 'info');\n"
        "    }\n"
        "  }\n"
        "\n"
        "  const sorted = deviceReports();\n"
        "  sorted.forEach(function(report){\n"
        "    const option = document.createElement('option');\n"
        "    option.value = report.date;\n"
        "    option.textContent = report.date;\n"
        "    select.appendChild(option);\n"
        "  });\n"
        "  if (!sorted.length) {\n"
        "    document.getElementById('report-content').innerHTML = '<p class=\"text-slate-500 text-sm text-center py-16\">No reports on this device yet.<br>Published data will sync automatically when available.</p>';\n"
        "  } else {\n"
        "    select.value = sorted[0].date;\n"
        "    setSelection(sorted[0]);\n"
        "  }\n"
        "  select.addEventListener('change', function(){\n"
        "    const selected = reports.find(function(r){ return r.date === select.value; });\n"
        "    if (selected) setSelection(selected);\n"
        "  });\n"
        "  initialPublishedSync();\n"
        "  maybeAutoCheckPublishedReports();\n"
        "}\n"
        "document.addEventListener('DOMContentLoaded', initIndex);\n"
        "</script>\n"
        "</body></html>"
    )


def write_index_page(docs_dir: Path):
    # Keep all already published reports visible in the index.
    # Publication itself is controlled by synchronized snapshots in main(),
    # so filtering by current raw files here can accidentally hide valid history.
    reports = collect_report_files(docs_dir)

    html_path = docs_dir / "index.html"
    trigger_api_url = os.getenv("REPORT_TRIGGER_API_URL", "").strip()
    html_path.write_text(format_report_index(reports, trigger_api_url), encoding="utf-8")

    # Write server-side last run timestamp so the page can show it regardless of browser state
    last_run_path = docs_dir / "last_run.json"
    last_run_path.write_text(
        json.dumps({"lastRunAt": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")}),
        encoding="utf-8",
    )


def parse_source_and_date(filename: str):
    """Extract source name and date from supported filename patterns."""
    date_match = DATE_RE.search(filename)
    file_date = date_match.group(1) if date_match else None

    if "_data_" in filename:
        source_name = filename.split("_data_")[0]
    else:
        # Legacy pattern: source_YYYY-MM-DD.xlsx
        source_name = re.sub(r"_\d{4}-\d{2}-\d{2}\.xlsx$", "", filename)

    return source_name, file_date


def parse_iso_date(value: str):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def institution_from_source(source_name: str) -> str:
    """Return the institution prefix from a source name (e.g. 'swedbank_pensions' -> 'swedbank')."""
    return source_name.split("_")[0]


def normalize_for_matching(text: str) -> str:
    return (
        str(text)
        .lower()
        .replace("į", "i")
        .replace("š", "s")
        .replace("ų", "u")
        .replace("ū", "u")
        .replace("ą", "a")
        .replace("č", "c")
        .replace("ę", "e")
        .replace("ė", "e")
        .replace("ž", "z")
    )


def fund_bucket_order(fund_name: str) -> int:
    name = normalize_for_matching(fund_name)

    for bucket, order in FUND_BUCKET_MAP.items():
        if bucket in name:
            return order

    if "turto issaugojimo" in name or "turto isaugojimo" in name:
        return len(FUND_AGE_BUCKETS)

    return len(FUND_AGE_BUCKETS) + 1


def discover_latest_files_per_source():
    """
    Discover latest synchronized Excel snapshot where all providers have data.

    Rules:
    - Parse all provider source files from both naming styles.
    - Compute latest date present for every provider in PROVIDER_ORDER.
    - Return files only for that date, so partial provider uploads do not advance reports.
    """
    candidates = []
    for path in Path(".").glob("*.xlsx"):
        lower = path.name.lower()
        if path.name.startswith("~$"):
            continue
        if "combined" in lower:
            continue
        if not DATE_RE.search(path.name):
            continue
        candidates.append(path)

    # source -> {date_str -> best_file_for_that_source_and_date}
    files_by_source_and_date = {}
    # institution -> set(date_str)
    dates_by_institution = {}

    for path in candidates:
        source_name, file_date = parse_source_and_date(path.name)
        if not source_name or not file_date:
            continue

        institution = institution_from_source(source_name)
        files_by_source_and_date.setdefault(source_name, {})
        dates_by_institution.setdefault(institution, set()).add(file_date)

        existing = files_by_source_and_date[source_name].get(file_date)
        if existing is None or path.stat().st_mtime > existing.stat().st_mtime:
            files_by_source_and_date[source_name][file_date] = path

    required_providers = set(PROVIDER_ORDER)
    missing_providers = sorted(required_providers - set(dates_by_institution.keys()))
    if missing_providers:
        raise RuntimeError(
            "Missing provider files for: " + ", ".join(missing_providers)
        )

    common_dates = None
    for provider in PROVIDER_ORDER:
        provider_dates = dates_by_institution.get(provider, set())
        common_dates = provider_dates if common_dates is None else (common_dates & provider_dates)

    if not common_dates:
        raise RuntimeError(
            "No synchronized date available across all providers yet."
        )

    selected_date = max(common_dates)

    by_source = {}
    for source_name, dated_files in files_by_source_and_date.items():
        picked = dated_files.get(selected_date)
        if picked is not None:
            by_source[source_name] = picked

    selected_institutions = {
        institution_from_source(source_name)
        for source_name in by_source.keys()
    }
    still_missing = sorted(required_providers - selected_institutions)
    if still_missing:
        raise RuntimeError(
            "Latest synchronized date is incomplete for providers: " + ", ".join(still_missing)
        )

    return by_source


def main():
    print("Discovering data files...")
    try:
        data_files = discover_latest_files_per_source()
    except RuntimeError as exc:
        docs_dir = Path("docs")
        docs_dir.mkdir(exist_ok=True)
        existing_reports = collect_report_files(docs_dir)

        if existing_reports:
            # Keep workflow successful and preserve all already published reports
            # when fresh provider files are temporarily out-of-sync.
            print(f"Warning: {exc}")
            latest_existing = existing_reports[-1]["date"]
            print(f"No synchronized snapshot available yet. Keeping published report date: {latest_existing}")
            write_index_page(docs_dir)
            print("✅ Index page refreshed while preserving existing published reports.")
            return

        raise

    if not data_files:
        print("Error: No data files found. Run scrapers first.")
        sys.exit(1)

    print(f"Found {len(data_files)} data source(s):")
    for source, filepath in sorted(data_files.items()):
        print(f"  - {source}: {filepath.name}")

    snapshot_dates = sorted(
        {
            parse_source_and_date(filepath.name)[1]
            for filepath in data_files.values()
            if parse_source_and_date(filepath.name)[1]
        }
    )
    if snapshot_dates:
        print(f"Using synchronized snapshot date: {snapshot_dates[-1]}")

    # Read all source files, grouped by institution.
    by_institution = {}
    for source, filepath in sorted(data_files.items()):
        print(f"\nReading {source}...")
        df = pd.read_excel(filepath)
        print(f"  Loaded {len(df)} records, {len(df.columns)} columns")

        institution = institution_from_source(source)
        _, file_date = parse_source_and_date(filepath.name)

        if institution not in by_institution:
            by_institution[institution] = {"file_date": file_date, "dfs": []}
        by_institution[institution]["dfs"].append(df)

    # Within each institution, merge all its files on Fund name so each fund is one row.
    # Across institutions, stack rows.
    institution_frames = []
    for institution, info in sorted(by_institution.items()):
        dfs = info["dfs"]

        if len(dfs) == 1:
            merged = dfs[0].copy()
        else:
            merged = dfs[0]
            for other in dfs[1:]:
                merged = merged.merge(other, on="Fund name", how="outer")

        merged["_institution"] = institution

        institution_frames.append(merged)
        print(f"  Institution '{institution}': {len(merged)} funds")

    print("\nCombining all institutions into one table...")
    df_combined = pd.concat(institution_frames, ignore_index=True, sort=False)

    # Exclude closed legacy funds that should no longer appear in reports.
    if "Fund name" in df_combined.columns:
        df_combined = df_combined[~df_combined["Fund name"].astype(str).str.contains(r"1954-1960|54/60", case=False, regex=True)]

    # Consolidate equivalent columns from different sources:
    # Date (Swedbank) -> Data
    if "Date" in df_combined.columns:
        if "Data" not in df_combined.columns:
            df_combined["Data"] = df_combined["Date"]
        else:
            df_combined["Data"] = df_combined["Data"].combine_first(df_combined["Date"])
        df_combined.drop(columns=["Date"], inplace=True)

    # GAV (Swedbank) -> Vieneto vertė
    if "GAV" in df_combined.columns:
        if "Vieneto vertė" not in df_combined.columns:
            df_combined["Vieneto vertė"] = df_combined["GAV"]
        else:
            df_combined["Vieneto vertė"] = df_combined["Vieneto vertė"].combine_first(df_combined["GAV"])
        df_combined.drop(columns=["GAV"], inplace=True)

    # Fondo dydis value (Swedbank) -> Grynieji aktyvai
    if "Fondo dydis value" in df_combined.columns:
        if "Grynieji aktyvai" not in df_combined.columns:
            df_combined["Grynieji aktyvai"] = df_combined["Fondo dydis value"]
        else:
            df_combined["Grynieji aktyvai"] = df_combined["Grynieji aktyvai"].combine_first(df_combined["Fondo dydis value"])
        df_combined.drop(columns=["Fondo dydis value"], inplace=True)

    # Group equivalent funds together: age bucket first, then provider.
    if "Fund name" in df_combined.columns:
        df_combined["_bucket_order"] = df_combined["Fund name"].apply(fund_bucket_order)
        df_combined["_provider_order"] = (
            df_combined.get("_institution", "")
            .astype(str)
            .map(PROVIDER_ORDER_MAP)
            .fillna(len(PROVIDER_ORDER))
        )
        df_combined.sort_values(
            ["_provider_order", "_bucket_order", "Fund name"],
            ignore_index=True,
            inplace=True,
        )

        # Insert a provider header row before each provider block.
        if "_institution" in df_combined.columns:
            block_rows = []
            current_provider = None
            for _, row in df_combined.iterrows():
                provider = str(row.get("_institution", ""))
                if provider != current_provider:
                    current_provider = provider
                    separator = {col: "" for col in df_combined.columns}
                    separator["Fund name"] = f"{provider.upper()}"
                    block_rows.append(separator)
                block_rows.append(row.to_dict())
            df_combined = pd.DataFrame(block_rows)

        df_combined.drop(columns=["_bucket_order", "_provider_order", "_institution"], inplace=True, errors="ignore")

    # Normalise Data column to YYYY-MM-DD (replace spaces/slashes with dashes)
    if "Data" in df_combined.columns:
        df_combined["Data"] = (
            df_combined["Data"]
            .astype(str)
            .str.strip()
            .str.replace(r"[\s/.]", "-", regex=True)
        )

    def clean_numeric(series):
        return pd.to_numeric(
            series.astype(str)
            .str.replace("EUR", "", regex=False)
            .str.replace(r"\s", "", regex=True)   # remove all whitespace (thousands sep)
            .str.replace(",", ".", regex=False)    # normalise decimal comma → dot
            .str.strip(),
            errors="coerce"
        )

    # Clean Vieneto vertė: strip "EUR", convert to numeric
    if "Vieneto vertė" in df_combined.columns:
        df_combined["Vieneto vertė"] = clean_numeric(df_combined["Vieneto vertė"])

    # Clean Grynieji aktyvai: strip "EUR", remove space thousands sep, convert to numeric
    if "Grynieji aktyvai" in df_combined.columns:
        df_combined["Grynieji aktyvai"] = clean_numeric(df_combined["Grynieji aktyvai"])

    print(f"  Combined: {len(df_combined)} rows, {len(df_combined.columns)} columns")

    # Use latest valid date from combined data for filename; fallback to today.
    if 'Data' in df_combined.columns:
        normalized_dates = (
            df_combined['Data']
            .dropna()
            .astype(str)
            .str.strip()
            .str.replace(r"[\s/.]", "-", regex=True)
            .str.extract(r"(\d{4}-\d{2}-\d{2})", expand=False)
            .dropna()
            .tolist()
        )
        data_date = max(normalized_dates) if normalized_dates else datetime.today().strftime("%Y-%m-%d")
    else:
        data_date = datetime.today().strftime("%Y-%m-%d")

    output_file = f"pension_data_combined_{data_date}.xlsx"

    # Rename column before writing
    df_combined.rename(columns={"Fund name": "Fondo pavadinimas"}, inplace=True)

    print(f"\nWriting to {output_file}...")
    df_combined.to_excel(output_file, index=False)

    # Apply formatting
    wb = load_workbook(output_file)
    ws = wb.active

    # Column widths
    ws.column_dimensions["A"].width = 40.12
    for col_letter in ["B", "C", "D"]:
        ws.column_dimensions[col_letter].width = 21.5

    # Header row: bold, size 14, centered
    header_font = Font(bold=True, size=14)
    header_align = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_align

    wb.save(output_file)

    # Also write an HTML report into docs/ so GitHub Pages can serve it
    try:
        docs_dir = Path("docs")
        docs_dir.mkdir(exist_ok=True)
        html_path = docs_dir / f"pension_data_combined_{data_date}.html"
        # Copy the Excel file into docs so it can be downloaded from GitHub Pages
        try:
            shutil.copy(output_file, docs_dir / output_file)
        except Exception:
            pass
        # Use a simple styled wrapper for readability
        # Prepare a display copy: replace NaN with empty string and format numbers
        display_df = df_combined.copy()
        display_df = display_df.fillna("")

        def fmt_gross(x):
            try:
                if x == "":
                    return ""
                return f"{int(round(float(x))):,}"
            except Exception:
                return x

        def fmt_unit(x):
            try:
                if x == "":
                    return ""
                return f"{float(x):,.4f}"
            except Exception:
                return x

        if "Grynieji aktyvai" in display_df.columns:
            display_df["Grynieji aktyvai"] = display_df["Grynieji aktyvai"].apply(fmt_gross)
        if "Vieneto vertė" in display_df.columns:
            display_df["Vieneto vertė"] = display_df["Vieneto vertė"].apply(fmt_unit)

        html_table = display_df.to_html(index=False, escape=False)
        # Replace provider header rows (e.g. ALLIANZ) with a full-width provider row
        try:
            ncols = len(display_df.columns)
            for prov in PROVIDER_ORDER:
                prov_up = prov.upper()
                # pattern: a row where first td == prov_up and remaining tds are empty
                pattern = rf"<tr>\s*<td[^>]*>{prov_up}</td>(?:\s*<td[^>]*>\s*</td>){{{ncols-1}}}\s*</tr>"
                replacement = f"<tr class=\"provider\"><td colspan=\"{ncols}\">{prov_up}</td></tr>"
                html_table = re.sub(pattern, replacement, html_table, flags=re.IGNORECASE)
        except Exception:
            pass

        # Add Tailwind styling and a cleaner report page layout
        html_table = html_table.replace(
            '<table border="1" class="dataframe">',
            '<table border="0" class="dataframe min-w-full divide-y divide-slate-200 text-sm text-slate-700 bg-white shadow-sm">'
        )
        html_table = html_table.replace('<thead>', '<thead class="bg-slate-100 text-slate-900">')
        html_table = html_table.replace('<th>', '<th class="whitespace-nowrap px-3 py-2 sm:px-4 sm:py-3 text-left font-semibold text-slate-900">')
        html_table = html_table.replace('<td>', '<td class="px-3 py-2 sm:px-4 sm:py-3">')
        html_table = re.sub(
            r'<tr class="provider"><td colspan="(\d+)">(.*?)</td></tr>',
            r'<tr class="provider bg-slate-100 text-slate-700 uppercase tracking-[0.15em]"><td colspan="\1" class="px-4 py-3 font-semibold">\2</td></tr>',
            html_table,
            flags=re.IGNORECASE,
        )

        trigger_api_url = os.getenv("REPORT_TRIGGER_API_URL", "").strip()

        html_content = (
            "<!doctype html>\n"
            "<html lang=\"en\"><head><meta charset=\"utf-8\"><meta name=\"viewport\" content=\"width=device-width,initial-scale=1,viewport-fit=cover\">\n"
            f"<title>Pension data {data_date}</title>\n"
            "<script src=\"https://cdn.tailwindcss.com\"></script>\n"
            "<style>\n"
            "  table.dataframe { border-collapse: collapse; }\n"
            "  table.dataframe td, table.dataframe th { border-bottom: 1px solid #e2e8f0; }\n"
            "  table.dataframe tbody tr:nth-child(even) { background: #f8fafc; }\n"
            "  table.dataframe tbody tr.provider td { background: #f1f5f9; }\n"
            "</style>\n"
            "</head><body class=\"bg-slate-50 text-slate-900 antialiased\">\n"
            "<div class=\"min-h-screen py-4 px-3 pb-[max(2rem,env(safe-area-inset-bottom))] sm:py-10 sm:px-6 lg:px-8\">\n"
            "  <main class=\"mx-auto max-w-6xl\">\n"
            "    <section class=\"overflow-hidden rounded-[24px] sm:rounded-[32px] border border-slate-200 bg-white/95 p-4 sm:p-6 shadow-[0_30px_80px_-40px_rgba(15,23,42,0.20)]\">\n"
            "      <div class=\"flex flex-col gap-4 md:flex-row md:items-start md:justify-between\">\n"
            "        <div class=\"min-w-0\">\n"
            f"          <p class=\"text-xs uppercase tracking-[0.35em] text-slate-500\">Report overview</p>\n"
            f"          <h1 class=\"mt-3 text-3xl font-semibold tracking-tight text-slate-950 sm:text-4xl\">Pension data {data_date}</h1>\n"
            f"          <p class=\"mt-2 text-sm text-slate-500\">Generated: {datetime.now().isoformat(timespec='seconds')}</p>\n"
            "        </div>\n"
            "        <div class=\"grid w-full sm:w-auto grid-cols-1 sm:grid-cols-3 items-center gap-2 sm:gap-3\">\n"
            "          <a class=\"inline-flex items-center justify-center rounded-full border border-slate-300 bg-slate-50 px-4 py-2.5 text-sm font-semibold text-slate-700 transition hover:bg-slate-100\" href=\"index.html\" target=\"_top\">← Back to history</a>\n"
            f"          <a class=\"inline-flex items-center justify-center rounded-full bg-slate-950 px-4 py-2.5 text-sm font-semibold text-white transition hover:bg-slate-800\" href=\"{output_file}\" download>Download Excel</a>\n"
            "          <button id=\"check-updates-btn\" class=\"inline-flex items-center justify-center rounded-full border border-slate-400 bg-white px-4 py-2.5 text-sm font-semibold text-slate-700 transition hover:bg-slate-100\" onclick=\"checkForUpdates()\">🔄 Check for updates</button>\n"
            "        </div>\n"
            "      </div>\n"
            "      <div class=\"mt-6 grid gap-4 sm:grid-cols-[minmax(0,1fr)_auto]\">\n"
            "        <input id=\"filter-input\" class=\"h-12 w-full rounded-3xl border border-slate-200 bg-slate-50 px-4 text-slate-900 shadow-sm outline-none transition focus:border-slate-500 focus:ring-2 focus:ring-slate-200\" placeholder=\"Filter funds…\" aria-label=\"Filter funds\" />\n"
            "        <p class=\"text-sm text-slate-500\">Filter the table by fund name, provider, or date.</p>\n"
            "      </div>\n"
            "      <div class=\"mt-6 -mx-1 overflow-x-auto rounded-3xl border border-slate-200 bg-slate-50 p-0 sm:mx-0\">\n"
            "        <div class=\"overflow-hidden rounded-3xl\">\n"
            + html_table +
            "        </div>\n"
            "      </div>\n"
            "    </section>\n"
            "  </main>\n"
            "</div>\n"
            "<script src=\"https://cdn.jsdelivr.net/npm/tablesort@5.2.1/dist/tablesort.min.js\"></script>\n"
            "<script>\n"
            f"const currentReportDate = '{data_date}';\n"
            f"const TRIGGER_API_URL = '{trigger_api_url}';\n"
            "function showNotification(message, type = 'info') {\n"
            "  const notification = document.createElement('div');\n"
            "  const bgClass = type === 'success' ? 'bg-green-50 border-green-200 text-green-800' :\n"
            "                  type === 'warning' ? 'bg-yellow-50 border-yellow-200 text-yellow-800' :\n"
            "                  type === 'error' ? 'bg-red-50 border-red-200 text-red-800' :\n"
            "                  'bg-blue-50 border-blue-200 text-blue-800';\n"
            "  notification.className = `fixed top-3 left-3 right-3 sm:left-auto sm:right-4 sm:max-w-sm p-4 rounded-2xl border ${bgClass} shadow-lg z-50`;\n"
            "  notification.textContent = message;\n"
            "  document.body.appendChild(notification);\n"
            "  setTimeout(() => notification.remove(), 7000);\n"
            "}\n"
            "function extractLatestDate(indexHtml) {\n"
            "  const dateMatches = indexHtml.match(/date:\\s*\"(\\d{4}-\\d{2}-\\d{2})\"/g);\n"
            "  if (!dateMatches || !dateMatches.length) {\n"
            "    return null;\n"
            "  }\n"
            "  return dateMatches[dateMatches.length - 1].match(/\\d{4}-\\d{2}-\\d{2}/)[0];\n"
            "}\n"
            "async function fetchLatestDate() {\n"
            "  const response = await fetch(`index.html?t=${Date.now()}`, { cache: 'no-store' });\n"
            "  const html = await response.text();\n"
            "  return extractLatestDate(html);\n"
            "}\n"
            "function getStatusUrl(afterIso) {\n"
            "  const base = TRIGGER_API_URL.endsWith('/trigger')\n"
            "    ? TRIGGER_API_URL.slice(0, -8) + '/status'\n"
            "    : TRIGGER_API_URL.replace(/\\/+$/, '') + '/status';\n"
            "  return `${base}?after=${encodeURIComponent(afterIso || '')}`;\n"
            "}\n"
            "async function triggerWorkflow() {\n"
            "  if (!TRIGGER_API_URL) {\n"
            "    showNotification('Auto trigger endpoint is not configured.', 'error');\n"
            "    return null;\n"
            "  }\n"
            "  if (window.location.protocol === 'file:') {\n"
            "    showNotification('Update trigger is unavailable from local file preview. Open the deployed page or run a local web server.', 'warning');\n"
            "    return null;\n"
            "  }\n"
            "  try {\n"
            "    const response = await fetch(TRIGGER_API_URL, {\n"
            "      method: 'POST',\n"
            "      headers: { 'Content-Type': 'application/json' },\n"
            "      body: JSON.stringify({ action: 'trigger_daily_publish' })\n"
            "    });\n"
            "    const payload = await response.json().catch(function(){ return {}; });\n"
            "    if (response.ok) {\n"
            "      return payload.dispatchedAt || new Date().toISOString();\n"
            "    }\n"
            "    const detail = payload && payload.error ? `: ${payload.error}` : '';\n"
            "    showNotification(`Trigger endpoint failed (${response.status})${detail}`, 'error');\n"
            "    return null;\n"
            "  } catch (error) {\n"
            "    let detail = error && error.message ? error.message : 'network error';\n"
            "    if (window.location.protocol === 'file:') {\n"
            "      detail = 'Request blocked in this context. Use the deployed page or local web server.';\n"
            "    } else if (detail === 'Failed to fetch') {\n"
            "      detail = 'Network error. Check your connection, or it may be a CORS restriction. Check browser console (F12) for details.';\n"
            "    }\n"
            "    showNotification(`Error calling trigger endpoint: ${detail}`, 'error');\n"
            "    return null;\n"
            "  }\n"
            "}\n"
            "async function waitForBuildCompletion(dispatchedAt, maxWaitMs = 420000) {\n"
            "  const startTime = Date.now();\n"
            "  const statusUrl = getStatusUrl(dispatchedAt);\n"
            "  while (Date.now() - startTime < maxWaitMs) {\n"
            "    await new Promise(resolve => setTimeout(resolve, 8000));\n"
            "    try {\n"
            "      const response = await fetch(`${statusUrl}&t=${Date.now()}`, { cache: 'no-store' });\n"
            "      if (!response.ok) {\n"
            "        continue;\n"
            "      }\n"
            "      const payload = await response.json().catch(function(){ return {}; });\n"
            "      if (payload.run && payload.run.status === 'completed') {\n"
            "        return payload.run;\n"
            "      }\n"
            "    } catch (e) {}\n"
            "  }\n"
            "  return null;\n"
            "}\n"
            "async function checkForUpdates() {\n"
            "  const btn = document.getElementById('check-updates-btn');\n"
            "  const defaultLabel = 'Check for updates';\n"
            "  btn.disabled = true;\n"
            "  btn.textContent = 'Triggering update...';\n"
            "  try {\n"
            "    const beforeDate = (await fetchLatestDate()) || currentReportDate;\n"
            "    const dispatchedAt = await triggerWorkflow();\n"
            "    if (!dispatchedAt) {\n"
            "      return;\n"
            "    }\n"
            "    showNotification('Workflow triggered. Waiting for build to finish...', 'info');\n"
            "    btn.textContent = 'Waiting for build...';\n"
            "    const completedRun = await waitForBuildCompletion(dispatchedAt);\n"
            "    if (!completedRun) {\n"
            "      showNotification('Build did not finish in time. Please retry in a moment.', 'warning');\n"
            "      return;\n"
            "    }\n"
            "    sessionStorage.setItem('postBuildBaselineDate', beforeDate);\n"
            "    window.location.reload();\n"
            "  } catch (error) {\n"
            "    showNotification(`Update check failed: ${error.message}`, 'error');\n"
            "  } finally {\n"
            "    btn.disabled = false;\n"
            "    btn.textContent = defaultLabel;\n"
            "  }\n"
            "}\n"
            "document.addEventListener('DOMContentLoaded', function(){\n"
            "  var previousDate = sessionStorage.getItem('postBuildBaselineDate');\n"
            "  if (previousDate) {\n"
            "    sessionStorage.removeItem('postBuildBaselineDate');\n"
            "    fetchLatestDate().then(function(latestDate){\n"
            "      if (latestDate && latestDate > previousDate) {\n"
            "        showNotification(`New data found (${latestDate}).`, 'success');\n"
            "      } else {\n"
            "        showNotification('Latest data is already there (no newer report after pipeline run).', 'warning');\n"
            "      }\n"
            "    }).catch(function(){});\n"
            "  }\n"
            "  var table = document.querySelector('table.dataframe');\n"
            "  if(table){ try{ new Tablesort(table); }catch(e){} }\n"
            "  var input = document.getElementById('filter-input');\n"
            "  if(input && table){\n"
            "    input.addEventListener('input', function(){\n"
            "      var q = this.value.toLowerCase();\n"
            "      var rows = table.tBodies[0].rows;\n"
            "      var currentProvider = '';\n"
            "      var providerMatches = false;\n"
            "      for (var i=0;i<rows.length;i++){\n"
            "        var r = rows[i];\n"
            "        if(r.classList.contains('provider')){\n"
            "          currentProvider = r.textContent.toLowerCase().trim();\n"
            "          providerMatches = q && currentProvider.indexOf(q) > -1;\n"
            "          r.style.display = q ? (providerMatches ? '' : 'none') : '';\n"
            "          continue;\n"
            "        }\n"
            "        var text = r.textContent.toLowerCase();\n"
            "        var matches = q && (text.indexOf(q) > -1 || providerMatches);\n"
            "        r.style.display = q ? (matches ? '' : 'none') : '';\n"
            "      }\n"
            "    });\n"
            "  }\n"
            "});\n"
            "</script>\n"
            "</body></html>"
        )
        html_path.write_text(html_content, encoding="utf-8")
        # Keep full published history; do not prune old report files.
        write_index_page(docs_dir)
        print(f"\n✅ HTML report written to: {html_path}")
    except Exception as _e:
        print(f"Warning: failed to write HTML report: {_e}")

    print(f"\n✅ Merged file created: {output_file}")
    print(f"   Rows: {len(df_combined)}")
    print(f"   Columns: {list(df_combined.columns)}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"Error: {exc}")
        sys.exit(1)
